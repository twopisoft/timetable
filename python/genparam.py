from openpyxl import Workbook, load_workbook
from consts import SLOTS, DAYS, WK_SLOTS, WEEK_HRS, MIN_HRS, MAX_HRS, START_ROW, END_ROW, DAY_START_COL, DAY_END_COL, \
                   GROUP_COL, COURSE_CODE_COL, COURSE_NAME_COL, WEEKDAYS, TEACHER_RANGE, MAX_TEACHERS, MIN_HRSD, MAX_HRSD
import random
import math
import sys
import warnings
import argparse
import logging
import re

def __get_args():
    parser = argparse.ArgumentParser(prog='genparam.py', formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    
    parser.add_argument("schedule_xlsx", help="Input Schedule xlsx filename")
    parser.add_argument("teacher_xlsx", help="Input Teacher xlsx filename")
    parser.add_argument("param_file", help="Ouput param filename")

    parser.add_argument("--start_row", type=int, default=START_ROW, help="Start Row number in the Schedule xlsx file")
    parser.add_argument("--end_row", type=int, default=END_ROW, help="End Row number in the Schedule xlsx file")
    parser.add_argument("--day_start_col", default=DAY_START_COL, help="Start Column for day in input schedule xlsx file")
    parser.add_argument("--day_end_col", default=DAY_END_COL, help="End Column for day in input schedule xlsx file")

    parser.add_argument("--loglevel", default='info', choices=['info','warning'], help="Set the logging level")

    return parser.parse_args()

def __flatten(l):
    return [j for i in l for j in i]

def read_slots(fname,day_range,read_rooms=False):

    logging.info('Reading Slot Data ...')

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(filename=fname, read_only=True)

    ws = wb.active

    all_slots = []
    all_rooms = []
    re_room = re.compile(r"(\([\w\s*/]+\))",flags=re.UNICODE)

    for (i,row) in enumerate(ws.iter_rows(day_range)):
        group_slots = []
        group_room_slots = []
        for cell in row:
            day_slots = [0]*SLOTS
            room_slots = ['']*SLOTS
            slots = cell.value.split(u',')
            cur_room = ''
            for slot in slots[::-1]:
                s = slot.split(u'(')
                if len(s) > 1: 
                    cur_room = re_room.findall(slot)[0]
                #index = int(math.floor((int(s[0])-1)/2))
                index = int(s[0])-1
                day_slots[index] = 1
                room_slots[index] = cur_room
                
            group_slots.append(day_slots)
            group_room_slots.append(room_slots)

        all_slots.append(__flatten(group_slots))
        all_rooms.append(__flatten(group_room_slots))

    if read_rooms:
        return all_slots,all_rooms

    return all_slots

def read_teachers(fname, read_names=False):

    logging.info('Reading Teacher Data ...')

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(filename=fname, read_only=True)

    ws = wb.active

    availability = []

    '''
    rows = list(ws.rows)
    num_ent = 3 if read_names else 2
    for row in rows[1:]:
        hrs = [0]*num_ent
        r = list(row)
        start_index = 0 if read_names else 1
        print read_names,num_ent,start_index
        for (j,cell) in enumerate(r[start_index:]):
            value = cell.value
            print j,value
            hrs[j] = value if type(value) is unicode else int(value)
        availability.append(hrs)
    '''
    num_ent = 3 if read_names else 2
    num_special = 0
    for row in ws.iter_rows(TEACHER_RANGE):
        hrs = [0]*num_ent
        r = list(row)
        start_index = 0 if read_names else 1
        for (j,cell) in enumerate(r[start_index:]):
            value = cell.value
            hrs[j] = value if type(value) is unicode else int(value)

        # check max hours
        if hrs[num_ent-1] < WEEK_HRS:
            num_special += 1

        availability.append(hrs)

    return availability,num_special

def __availability():
    logging.info('Generating Availability data ...')

    special = random.sample(xrange(NUM_TEACHERS),int(math.ceil(NUM_TEACHERS*SPECIAL_PERCENT)))
    availabilty = [[MIN_HRS,MAX_HRS]]*NUM_TEACHERS
    for s in special:
        availabilty[s] = [MIN_HRS,MIN_HRS]

    return availabilty

def write_param_file(param_file,all_slots,availability,num_special):

    logging.info('Writing param file "{0}" ...'.format(param_file))
    with open(param_file,'w') as outfile:
        outfile.write('letting NUM_GROUPS     = {}\n'.format(len(all_slots)))
        outfile.write('letting NUM_TEACHERS   = {}\n'.format(len(availability)))
        outfile.write('letting NUM_SPECIAL    = {}\n'.format(num_special))
        outfile.write('letting MIN_HRS        = {}\n'.format(MIN_HRS))
        outfile.write('letting MAX_HRS        = {}\n'.format(MAX_HRS))
        outfile.write('letting WEEK_HRS       = {}\n'.format(WEEK_HRS))
        outfile.write('letting MAX_TEACHERS   = {}\n'.format(MAX_TEACHERS))
        outfile.write('letting MIND_HRS       = {}\n'.format(MIN_HRSD))
        outfile.write('letting MAXD_HRS       = {}\n'.format(MAX_HRSD))
        outfile.write('letting DAYS           = {}\n'.format(DAYS))
        outfile.write('letting DAY_SLOTS      = {}\n'.format(SLOTS))
        outfile.write('letting WEEK_SLOTS     = {}\n'.format(WK_SLOTS))
        outfile.write('\nletting Demand       = ' + str(all_slots))
        outfile.write('\n\nletting Availability = ' + str(availability))

def test_resource_match(all_slots, num_teachers) :

    data_ok = True

    for s in range(SLOTS*DAYS):
        busy_slots = 0
        for (i,g) in enumerate(all_slots):
            busy_slots += all_slots[i][s]

        if busy_slots > num_teachers:
            data_ok = False
            logging.warning('Found resource mismatch on {0} for slot {1}. busy_slots={2}, num_teachers={3}'.format(WEEKDAYS[s/SLOTS],s,busy_slots,num_teachers))

    if data_ok:
        logging.info('Resource match check: Ok')

    return data_ok

def test_week_hrs(all_slots):

    data_ok = True

    for (i,s) in enumerate(all_slots):
        hrs = s.count(1)
        if hrs != WEEK_HRS:
            data_ok = False
            logging.warning("Not enough hours for group {0}. {1} != {2}".format(i+1,hrs,WEEK_HRS))

    if data_ok:
        logging.info("Hours Scheduled: Ok")

    return data_ok

def main():
    args = __get_args()

    log_level = getattr(logging, args.loglevel.upper(), None)
    logging.basicConfig(level=log_level, format='%(asctime)s - %(levelname)s : %(message)s', datefmt='%Y-%m-%d %I:%M:%S %p')

    schedule_xlsx = args.schedule_xlsx
    teacher_xlsx = args.teacher_xlsx
    param_file = args.param_file

    start_row = args.start_row 
    end_row = args.end_row 

    num_groups = abs(start_row - end_row) + 1
    day_start_col = args.day_start_col 
    day_end_col = args.day_end_col

    day_range = '{0}{1}:{2}{3}'.format(day_start_col,start_row,day_end_col,end_row)

    all_slots = read_slots(schedule_xlsx,day_range)
    availability,num_special = read_teachers(teacher_xlsx)
    num_teachers = len(availability)

    test_resource_match(all_slots,num_teachers)
    test_week_hrs(all_slots)

    write_param_file(param_file,all_slots,availability,num_special)

if __name__ == "__main__":
    main()
