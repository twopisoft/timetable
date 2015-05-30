# coding=utf-8
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Style
import re

in_wb = load_workbook(filename='../data/first_term_schedule.xlsx', read_only=True)
out_wb = Workbook()
in_ws = in_wb.active
out_ws = out_wb.active

in_ws.sheet_view.rightToLeft=True
out_ws.sheet_view.rightToLeft=True

pat = r"([\(]*[\w\s*/]+\))"
align_style = Style(alignment=Alignment(wrap_text=True))
ro = 0
for (ri,irow) in enumerate(in_ws.iter_rows('A1:I320')):
    for (ci,icell) in enumerate(irow):
        if icell.value is not None:
            ivalue = icell.value
                
            if ci == 0:
                ro += 1
            
            ocell = out_ws.cell(row=ro, column=ci+1)
            val = ocell.value
            
            if val is not None:
                if type(ivalue) is unicode:
                    val = u"{0} {1}".format(val,ivalue)
            else:
                val = ivalue
            
            if type(val) is float:
                val = int(val)
            elif type(val) is unicode:
                val = val.strip()
                val = val.replace(u'–',u'')
                val = re.sub(pat, (lambda x: re.sub(r' ',u'',x.group(0))), val)
                val = re.sub(r"\s+",u' ',val)
                
            ocell.value = val
            ocell.style = align_style

for (r,orow) in enumerate(out_ws.iter_rows('E2:I100')):
    for (c,co) in enumerate(orow):
        co.value = co.value.replace(u' (',u':(').replace(u' ',u',').replace(u':',u' ')

out_wb.save('../data/first_term_schedule_cleaned2.xlsx')