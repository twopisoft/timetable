from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from genparam import read_teachers
from verify import parse_solution
from consts import SLOTS, DAYS, START_ROW, END_ROW, DAY_START_COL, DAY_END_COL, GROUP_COL, COURSE_CODE_COL, COURSE_NAME_COL, WEEKDAYS
import sys
import warnings
import argparse


def __get_args():
    parser = argparse.ArgumentParser()
    
    parser.add_argument("schedule_xlsx", help="Input Schedule xlsx filename")
    parser.add_argument("teacher_xlsx", help="Input Teacher xlsx filename")
    parser.add_argument("solution_file", help="Input solution filename")
    parser.add_argument("solved_xlsx", help="Output xlsx filename for student's timetable")

    parser.add_argument("--start_row", type=int, help="Start Row number in the Schedule xlsx file (default: {})".format(START_ROW))
    parser.add_argument("--end_row", type=int, help="End Row number in the Schedule xlsx file (default: {})".format(END_ROW))
    parser.add_argument("--day_start_col", help="Start Column for day in input schedule xlsx file (default: {})".format(DAY_START_COL))
    parser.add_argument("--day_end_col", help="End Column for day in input schedule xlsx file (default: {})".format(DAY_END_COL))
    parser.add_argument("--group_col", help="Group Column in input schedule xlsx file (default: {})".format(GROUP_COL))
    parser.add_argument("--course_code_col", help="Course Code Column in input schedule xlsx file (default: {})".format(COURSE_CODE_COL))
    parser.add_argument("--course_name_col", help="Course Name Column in input schedule xlsx file (default: {})".format(COURSE_NAME_COL))
    parser.add_argument("--solved_teacher_xlsx", help="Output xlsx filename for teacher's timetable")

    return parser.parse_args()

def copy_col_cells(in_ws, out_ws, in_data_range, out_col, has_header=True):

    out_col_index = column_index_from_string(out_col)
    row_offset = 2 if has_header else 1

    for (i,row) in enumerate(in_ws.iter_rows(in_data_range)):
        for cell in row:
            out_ws.cell(row=i+row_offset,column=out_col_index,value=cell.value)

    return out_ws

def copy_value(out_ws,value,data_range):
    
    for row in out_ws.iter_rows(data_range):
        for cell in row:
            cell.value = value

def copy_data(schedule_xlsx,start_row,num_rows,group_col,course_code_col,course_name_col):

    print 'Copying Data ...'

    out_wb = Workbook()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        in_wb = load_workbook(filename=schedule_xlsx, read_only=True)

    in_ws = in_wb.active
    out_ws = out_wb.active
    out_ws.title = in_ws.title

    header = ['Course Code','Course Name','Hours','Group'] + WEEKDAYS + ['Remarks']
    for (i,h) in enumerate(header):
        out_ws.cell(row=1, column=i+1, value=h)

    for (incol,outcol) in [(course_code_col,'A'),(course_name_col,'B'),(group_col,'D')]:
        out_ws = copy_col_cells(in_ws,out_ws,'{0}{1}:{0}{2}'.format(incol,start_row,start_row+num_rows-1),outcol)

    out_ws = copy_value(out_ws, 16, '{0}{1}:{0}{2}'.format('C',2,num_rows+1))

    return out_wb

def gen_student_tt(solved_xlsx, wb, teacher_data, roster):

    print 'Generating Student Timetable ...'

    ws = wb.active
    num_rows = len(roster)
    for row in range(2,num_rows+2):
        i = row - 2
        
        for d in range(DAYS):
            value = ''
            for s in range(SLOTS):
                t = roster[i][s+SLOTS*d]
                if t != 0:
                    value += '{0} - {1}'.format(s+1,teacher_data[t-1][0]) + '\n'
            ws.cell(row=row,column=5+d,value=value) 

    wb.save(filename=solved_xlsx)

def group_dict(schedule_xlsx,start_row,num_groups,group_col):

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(filename=schedule_xlsx, read_only=True)

    ws = wb.active

    g_range = '{0}{1}:{2}{3}'.format(group_col,start_row,group_col,(start_row+num_groups-1))
    gr_dict = {}

    for (g,row) in enumerate(ws.iter_rows(g_range)):
        for cell in row:
            gr_dict[g+1] = cell.value

    return gr_dict

def inverse_roster(roster, gr_dict):

    inv = {}

    for (g,group) in enumerate(roster):
        for (s,t) in enumerate(group):
            if t != 0:
                if not t in inv.keys():
                    inv[t] = [[0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0]]
                d = s/SLOTS
                st = int((s/float(SLOTS) - d)*SLOTS)
                inv[t][d][st] = gr_dict[g+1]

    return inv

def gen_teacher_tt(solved_teacher_xlsx, teacher_data, gr_dict, roster):

    print 'Generating Teacher Timetable ...'

    wb = Workbook()
    ws = wb.active

    header = ['Teacher Name'] + WEEKDAYS + ['Hours/Week']
    for (i,h) in enumerate(header):
        ws.cell(row=1, column=i+1, value=h)

    iroster = inverse_roster(roster, gr_dict)

    col_offset = 2
    row_offset = 2
    for (t,td) in enumerate(teacher_data):
        hrs = 0
        for d in range(DAYS):
            value = ''
            for s in range(SLOTS):
                g = iroster[t+1][d][s]
                if g != 0:
                    value += '{0} - {1}'.format(s+1,g) + '\n'
                    hrs += 1
            ws.cell(row=t+row_offset,column=d+col_offset,value=value)

        ws.cell(row=t+row_offset,column=1,value=td[0])
        ws.cell(row=t+row_offset,column=7,value=hrs)

    wb.save(solved_teacher_xlsx)

def main():
    args = __get_args()

    start_row = args.start_row or START_ROW
    end_row = args.end_row or  END_ROW
    group_col = args.group_col or GROUP_COL
    course_code_col = args.course_code_col or COURSE_CODE_COL
    course_name_col = args.course_name_col or COURSE_NAME_COL

    num_groups = abs(start_row - end_row) + 1
    day_start_col = args.day_start_col or DAY_START_COL
    day_end_col = args.day_end_col or DAY_END_COL

    day_range = '{0}{1}:{2}{3}'.format(day_start_col,start_row,day_end_col,end_row)

    solution_wb = copy_data(args.schedule_xlsx,start_row,num_groups,group_col,course_code_col,course_name_col)
    teacher_data = read_teachers(args.teacher_xlsx, read_names=True)
    roster = parse_solution(args.solution_file)

    gen_student_tt(args.solved_xlsx, solution_wb, teacher_data, roster)

    if args.solved_teacher_xlsx:
        gr_dict = group_dict(args.schedule_xlsx,start_row,num_groups,group_col)
        gen_teacher_tt(args.solved_teacher_xlsx,teacher_data, gr_dict, roster)

if __name__ == "__main__":
    main()