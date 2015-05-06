from openpyxl import Workbook
from consts import MIN_HRS, MAX_HRS
import random
import math
import sys
import warnings
import argparse

def __availability(num_teachers,special_percent):

    special = random.sample(xrange(num_teachers),int(math.ceil(num_teachers*special_percent)))
    availabilty = [[MIN_HRS,MAX_HRS]]*num_teachers
    for s in special:
        availabilty[s] = [MIN_HRS,MIN_HRS]

    return availabilty

parser = argparse.ArgumentParser()
parser.add_argument("teacher_xlsx", help="Output Teacher xlsx filename")
parser.add_argument("num_teachers", type=int, help="Number of teachers")
parser.add_argument("--special_percent", type=int, help="Percentage of teachers with limited hours (default: 20)")

args = parser.parse_args()

special_percent = (args.special_percent or 20)/100.0
num_teachers = args.num_teachers

wb = Workbook()

ws = wb.active
ws.title = 'Teachers'

ws['A1'] = 'Name'
ws['B1'] = 'Min Hours / week'
ws['C1'] = 'Max Hours / week'

availabilty = __availability(num_teachers,special_percent)

for i in range(1,num_teachers+1):
    ws.cell(column=1, row=i+1, value='T{0}'.format(i))
    ws.cell(column=2, row=i+1, value=availabilty[i-1][0])
    ws.cell(column=3, row=i+1, value=availabilty[i-1][1])

wb.save(filename=args.teacher_xlsx)
