from openpyxl import Workbook, load_workbook
from consts import SLOTS, DAYS, WEEK_HRS, MIN_HRS, MAX_HRS, START_ROW, END_ROW, DAY_START_COL, DAY_END_COL, GROUP_COL, COURSE_CODE_COL, COURSE_NAME_COL, WEEKDAYS
import random
import math
import sys
import warnings
import argparse

def __get_args():
    parser = argparse.ArgumentParser()
    
    parser.add_argument("schedule_xlsx", help="Input Schedule xlsx filename")
    parser.add_argument("teacher_xlsx", help="Input Teacher xlsx filename")
    parser.add_argument("param_file", help="Ouput param filename")

    parser.add_argument("--start_row", type=int, help="Start Row number in the Schedule xlsx file (default: 80)")
    parser.add_argument("--end_row", type=int, help="End Row number in the Schedule xlsx file (default: 165)")
    parser.add_argument("--day_start_col", help="Start Column for day in input schedule xlsx file (default: K)")
    parser.add_argument("--day_end_col", help="End Column for day in input schedule xlsx file (default: O)")

    return parser.parse_args()

def __flatten(l):
    return [j for i in l for j in i]

def read_slots(fname,day_range):

    print 'Reading Slot Data ...'

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(filename=fname, read_only=True)

    ws = wb.active

    all_slots = []
    for (i,row) in enumerate(ws.iter_rows(day_range)):
        group_slots = []
        for cell in row:
            day_slots = [0]*SLOTS
            slots = cell.value.split(',')
            for slot in slots:
                s = slot.split('(')
                #index = int(math.floor((int(s[0])-1)/2))
                index = int(s[0])-1
                day_slots[index] = 1
            group_slots.append(day_slots)
        all_slots.append(__flatten(group_slots))

    return all_slots

def read_teachers(fname, read_names=False):

    print 'Reading Teacher Data ...'

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(filename=fname, read_only=True)

    ws = wb.active

    availability = []
    rows = list(ws.rows)
    num_ent = 3 if read_names else 2
    for row in rows[1:]:
        hrs = [0]*num_ent
        r = list(row)
        start_index = 0 if read_names else 1
        for (j,cell) in enumerate(r[start_index:]):
            hrs[j] = cell.value
        availability.append(hrs)

    return availability

def __availability():
    print 'Generating Availability data ...'

    special = random.sample(xrange(NUM_TEACHERS),int(math.ceil(NUM_TEACHERS*SPECIAL_PERCENT)))
    availabilty = [[MIN_HRS,MAX_HRS]]*NUM_TEACHERS
    for s in special:
        availabilty[s] = [MIN_HRS,MIN_HRS]

    return availabilty

def write_param_file(param_file,all_slots,availability):

    print 'Writing param file "{0}" ...'.format(param_file)
    with open(param_file,'w') as outfile:
        outfile.write('letting NUM_GROUPS = {}\n'.format(len(all_slots)))
        outfile.write('letting NUM_TEACHERS = {}\n'.format(len(availability)))
        outfile.write('\nletting Demand = ' + str(all_slots))
        outfile.write('\n\nletting Availability = ' + str(availability))

def test_resource_match(all_slots, num_teachers) :

    data_ok = True

    for s in range(SLOTS*DAYS):
        busy_slots = 0
        for (i,g) in enumerate(all_slots):
            busy_slots += all_slots[i][s]

        if busy_slots > num_teachers:
            data_ok = False
            print 'Found resource mismatch on {0} for slot {1}. busy_slots={2}, num_teachers={3}'.format(WEEKDAYS[s/SLOTS],s,busy_slots,num_teachers)

    if data_ok:
        print 'Resource match check: Ok'

    return data_ok

def test_week_hrs(all_slots):

    data_ok = True

    for (i,s) in enumerate(all_slots):
        hrs = s.count(1)
        if hrs != WEEK_HRS:
            data_ok = False
            print "Not enough hours for group {0}. {1} != {2}".format(i+1,hrs,WEEK_HRS)

    if data_ok:
        print "Hours Scheduled: Ok"

    return data_ok

def main():
    args = __get_args()

    schedule_xlsx = args.schedule_xlsx
    teacher_xlsx = args.teacher_xlsx
    param_file = args.param_file

    start_row = args.start_row or START_ROW
    end_row = args.end_row or  END_ROW

    num_groups = abs(start_row - end_row) + 1
    day_start_col = args.day_start_col or DAY_START_COL
    day_end_col = args.day_end_col or DAY_END_COL

    day_range = '{0}{1}:{2}{3}'.format(day_start_col,start_row,day_end_col,end_row)

    all_slots = read_slots(schedule_xlsx,day_range)
    availability = read_teachers(teacher_xlsx)
    num_teachers = len(availability)

    test_resource_match(all_slots,num_teachers)
    test_week_hrs(all_slots)

    write_param_file(param_file,all_slots,availability)

if __name__ == "__main__":
    main()
